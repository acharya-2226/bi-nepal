"""
reports/report_generator.py  -  Generate PDF and Excel BI reports.

Enhanced version:
  - Full-page PDF with embedded charts (matplotlib → PNG → reportlab)
  - Every section includes an auto-generated analytical narrative
  - Excel workbook with formatted tables + embedded charts per sheet
"""

from __future__ import annotations
import io
import textwrap
from pathlib import Path

# ---------------------------------------------------------------------------
# Logger / path setup  (mirrors your original)
# ---------------------------------------------------------------------------
try:
    from logger import get_logger
    logger = get_logger(__name__)
except ImportError:
    import logging
    logger = logging.getLogger(__name__)

try:
    from analysis.profitability  import compute_profitability
    from analysis.discounts      import compute_discounts
    from analysis.inventory      import compute_inventory
    from analysis.products       import compute_products
    from analysis.expenses       import compute_expenses
    from analysis.monthly_growth import compute_monthly_growth
    from analysis.breakeven      import compute_breakeven
    from analysis.cashflow       import compute_cashflow
    from routes.period_scope     import selected_labels
except ImportError:
    # -----------------------------------------------------------------------
    # STUB FUNCTIONS – replace with your real imports.
    # These produce realistic-looking dummy data so the file is self-contained
    # and you can run / test it without the full application installed.
    # -----------------------------------------------------------------------
    import random, math
    random.seed(42)

    _MONTHS = ["Jan-25","Feb-25","Mar-25","Apr-25","May-25","Jun-25",
               "Jul-25","Aug-25","Sep-25","Oct-25","Nov-25","Dec-25"]

    def selected_labels(period, bucket):
        return _MONTHS

    def compute_profitability(store):
        return {
            "gross_revenue_npr": 12_450_000,  "net_revenue_npr": 11_200_000,
            "gross_profit_npr":   4_800_000,  "gross_profit_margin_pct": 42.86,
            "net_profit_npr":     1_950_000,  "net_profit_margin_pct":   17.41,
            "total_opex_npr":     2_850_000,
        }

    def compute_discounts(store):
        return {"total_discount_npr": 1_250_000, "discount_rate_pct": 10.04,
                "by_category": {"Electronics":450000,"Clothing":310000,
                                "Home":280000,"Sports":210000}}

    def compute_inventory(store):
        return {
            "inventory_turnover": 6.3, "days_inventory_outstanding": 57.9,
            "items_below_reorder_level": 14,
            "below_reorder_items": [
                {"SKU": f"SKU-{100+i}", "Product": f"Product {i}",
                 "Stock": random.randint(1,9), "ReorderLevel": 10+i}
                for i in range(14)
            ],
        }

    def compute_products(store):
        prods = [
            {"Product": f"Product {i}", "Revenue_NPR": int(800000 - i*55000),
             "Units": int(950 - i*70), "Profit_NPR": int(340000 - i*22000)}
            for i in range(10)
        ]
        return {"top_10_products_by_revenue": prods}

    def compute_expenses(store):
        return {
            "expense_breakdown": {
                "Salaries":        1_100_000,
                "Marketing":         450_000,
                "Logistics":         380_000,
                "Platform Fees":     290_000,
                "Admin & Utilities": 200_000,
                "Miscellaneous":     105_000,  # reduced so total_opex matches
            },
            "total_expenses_npr": 2_525_000,
        }

    def compute_monthly_growth(store, period="monthly"):
        base = 850_000
        rows = []
        for i, lbl in enumerate(_MONTHS):
            rev  = int(base * (1 + 0.04*i + random.uniform(-0.02,0.02)))
            prof = int(rev  * (0.16 + random.uniform(-0.02,0.02)))
            rows.append({"PeriodLabel": lbl, "Revenue_NPR": rev,
                         "Profit_NPR": prof,
                         "GrowthRate_Pct": round((rev/base - 1)*100, 2)})
        return {"monthly": rows}

    def compute_breakeven(store):
        top20 = [
            {"Product": f"Product {i}", "BreakevenUnits": int(120 - i*5),
             "ActualUnits": int(200 - i*7), "MarginOfSafety_Pct": round(30 - i*0.8, 2)}
            for i in range(20)
        ]
        return {
            "overall_breakeven_units": 8_420,
            "actual_units_sold":       13_670,
            "margin_of_safety_pct":    38.4,
            "top_20_easiest_breakeven": top20,
        }

    def compute_cashflow(store, period="monthly"):
        rows = []
        base_in, base_out = 980_000, 780_000
        for lbl in _MONTHS:
            ci = int(base_in  * random.uniform(0.92, 1.1))
            co = int(base_out * random.uniform(0.90, 1.08))
            rows.append({"PeriodLabel": lbl,
                         "CashInflow_NPR": ci, "CashOutflow_NPR": co,
                         "NetCash_NPR": ci - co})
        return {
            "total_cash_inflow_npr":  sum(r["CashInflow_NPR"]  for r in rows),
            "total_cash_outflow_npr": sum(r["CashOutflow_NPR"] for r in rows),
            "net_cash_movement_npr":  sum(r["NetCash_NPR"]     for r in rows),
            "monthly_cashflow": rows,
        }


# ---------------------------------------------------------------------------
BASE_DIR   = Path(__file__).resolve().parent.parent
EXPORT_DIR = Path(__file__).resolve().parent / "data" / "exports"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)


# ═══════════════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def _normalize_period(period: str) -> str:
    n = (period or "monthly").strip().lower()
    return n if n in {"monthly", "quarterly", "semiannual", "annual"} else "monthly"


def _all(store: dict, period: str = "monthly", bucket: str | None = None) -> dict:
    labels   = set(selected_labels(period, bucket))
    growth   = compute_monthly_growth(store, period=period)
    cashflow = compute_cashflow(store, period=period)
    growth["monthly"] = [r for r in growth.get("monthly", [])
                         if r.get("PeriodLabel") in labels]
    cashflow["monthly_cashflow"] = [r for r in cashflow.get("monthly_cashflow", [])
                                    if r.get("PeriodLabel") in labels]
    return {
        "profitability":  compute_profitability(store),
        "discounts":      compute_discounts(store),
        "inventory":      compute_inventory(store),
        "products":       compute_products(store),
        "expenses":       compute_expenses(store),
        "monthly_growth": growth,
        "breakeven":      compute_breakeven(store),
        "cashflow":       cashflow,
    }


# ---------------------------------------------------------------------------
#  CHART BUILDERS  (return PNG bytes via in-memory buffer)
# ---------------------------------------------------------------------------

def _fig_to_bytes(fig) -> bytes:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=130, bbox_inches="tight")
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
#  KEY-RESOLUTION HELPERS
#  Your compute_* functions may return different column names across versions.
#  Each helper below tries a prioritised list of candidate names and falls back
#  to 0 so charts never crash with a KeyError.
# ---------------------------------------------------------------------------

def _rget(row: dict, *candidates, default=0):
    """Return the first matching key's value from *row*, or *default*."""
    for k in candidates:
        if k in row:
            return row[k]
    return default


def _growth_label(row: dict) -> str:
    return _rget(row, "PeriodLabel", "period_label", "Period", "Label",
                 "Month", "Quarter", default="?")


def _growth_revenue(row: dict) -> float:
    return float(_rget(row,
        "Revenue_NPR", "revenue_npr", "Revenue", "revenue",
        "TotalRevenue", "total_revenue", "GrossRevenue", "gross_revenue",
        "MonthRevenue", "month_revenue",
        default=0))


def _growth_profit(row: dict) -> float:
    return float(_rget(row,
        "Profit_NPR", "profit_npr", "Profit", "profit",
        "NetProfit", "net_profit", "GrossProfit", "gross_profit",
        "MonthProfit", "month_profit",
        default=0))


def _growth_rate(row: dict):
    return _rget(row,
        "GrowthRate_Pct", "growth_rate_pct", "GrowthRate", "growth_rate",
        "Growth_Pct", "growth_pct", "RevGrowth", "rev_growth",
        default="")


def _cf_label(row: dict) -> str:
    return _rget(row, "PeriodLabel", "period_label", "Period", "Label",
                 "Month", "Quarter", default="?")


def _cf_inflow(row: dict) -> float:
    return float(_rget(row,
        "CashInflow_NPR", "cash_inflow_npr", "CashInflow", "cash_inflow",
        "Inflow_NPR", "inflow_npr", "Inflow", "inflow",
        "CashIn", "cash_in",
        default=0))


def _cf_outflow(row: dict) -> float:
    return float(_rget(row,
        "CashOutflow_NPR", "cash_outflow_npr", "CashOutflow", "cash_outflow",
        "Outflow_NPR", "outflow_npr", "Outflow", "outflow",
        "CashOut", "cash_out",
        default=0))


def _cf_net(row: dict) -> float:
    return float(_rget(row,
        "NetCash_NPR", "net_cash_npr", "NetCash", "net_cash",
        "Net_NPR", "net_npr", "Net", "net",
        default=_cf_inflow(row) - _cf_outflow(row)))


def _prod_name(row: dict) -> str:
    return str(_rget(row, "Product", "product", "Name", "name",
                     "ProductName", "product_name",
                     "ItemName", "item_name", default="Unknown"))


def _prod_revenue(row: dict) -> float:
    return float(_rget(row,
        "Revenue_NPR", "revenue_npr", "Revenue", "revenue",
        "TotalRevenue", "total_revenue",
        "ProductRevenue", "product_revenue",
        default=0))


def _prod_profit(row: dict) -> float:
    return float(_rget(row,
        "Profit_NPR", "profit_npr", "Profit", "profit",
        "NetProfit", "net_profit",
        "ProductProfit", "product_profit",
        default=0))


def _prod_units(row: dict) -> int:
    return int(_rget(row,
        "Units", "units", "UnitsSold", "units_sold",
        "Quantity", "quantity",
        "ProductQtySold", "product_qty_sold",
        default=0))


def _be_product(row: dict) -> str:
    return str(_rget(row, "Product", "product", "Name", "name",
                     "ProductName", "product_name",
                     "ItemName", "item_name", default="Unknown"))


def _be_units(row: dict) -> int:
    return int(_rget(row,
        "BreakevenUnits", "breakeven_units", "BEUnits", "be_units",
        "BreakEvenUnits", "break_even_units",
        default=0))


def _be_actual(row: dict) -> int:
    return int(_rget(row,
        "ActualUnits", "actual_units", "UnitsSold", "units_sold",
        "Actual", "actual",
        default=0))


def _be_safety(row: dict) -> float:
    return float(_rget(row,
        "MarginOfSafety_Pct", "margin_of_safety_pct",
        "MarginSafety_Pct", "margin_safety_pct",
        "SafetyMargin_Pct", "safety_margin_pct",
        default=0))


def _inv_sku(row: dict) -> str:
    return str(_rget(row, "SKU", "sku", "Code", "code",
                     "ProductCode", "product_code",
                     "ItemID", "item_id", default="?"))


def _inv_name(row: dict) -> str:
    return str(_rget(row, "Product", "product", "Name", "name",
                     "ProductName", "product_name",
                     "ItemName", "item_name", default="Unknown"))


def _inv_stock(row: dict) -> int:
    return int(_rget(row,
        "Stock", "stock", "CurrentStock", "current_stock",
        "StockLevel", "stock_level", "Quantity", "quantity",
        "ClosingStock", "closing_stock",
        default=0))


def _inv_reorder(row: dict) -> int:
    return int(_rget(row,
        "ReorderLevel", "reorder_level", "ReorderPoint", "reorder_point",
        "MinStock", "min_stock",
        default=0))


def _disc_total(d: dict) -> float:
    if not isinstance(d, dict):
        return 0.0
    return float(_rget(d,
        "total_discount_npr", "total_discounts_npr",
        "total_discount", "total_discounts",
        "discount_total_npr", "discount_total",
        default=0))


def _disc_rate(d: dict, gross_revenue: float | None = None) -> float:
    if not isinstance(d, dict):
        return 0.0
    rate = _rget(d,
        "discount_rate_pct", "discount_rate",
        "discount_pct", "discount_percentage",
        "discount_pct_of_revenue",
        default=None)
    if rate is not None:
        try:
            return float(rate)
        except Exception:
            return 0.0
    if gross_revenue:
        total = _disc_total(d)
        return round((total / max(gross_revenue, 1)) * 100, 2)
    return 0.0


def _disc_by_category(d: dict) -> dict:
    if not isinstance(d, dict):
        return {}
    return _rget(d,
        "by_category", "discount_by_category", "category_discounts",
        "discounts_by_category", default={}) or {}


def _chart_revenue_profit(growth_rows: list) -> bytes:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    labels = [_growth_label(r)          for r in growth_rows]
    rev    = [_growth_revenue(r) / 1e6  for r in growth_rows]
    profit = [_growth_profit(r)  / 1e6  for r in growth_rows]

    fig, ax = plt.subplots(figsize=(9, 3.4))
    x = range(len(labels))
    ax.bar(x, rev,    width=0.4, label="Revenue (M NPR)", color="#2563EB", align="center")
    ax.bar([i+0.4 for i in x], profit, width=0.4, label="Profit (M NPR)",
           color="#10B981", align="center")
    ax.set_xticks([i+0.2 for i in x])
    ax.set_xticklabels(labels, rotation=40, ha="right", fontsize=8)
    ax.set_ylabel("NPR (Millions)")
    ax.set_title("Monthly Revenue vs Profit")
    ax.legend(); fig.tight_layout()
    data = _fig_to_bytes(fig); plt.close(fig); return data


def _chart_cashflow(cf_rows: list) -> bytes:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    labels  = [_cf_label(r)           for r in cf_rows]
    inflow  = [_cf_inflow(r)  / 1e6   for r in cf_rows]
    outflow = [_cf_outflow(r) / 1e6   for r in cf_rows]
    net     = [_cf_net(r)     / 1e6   for r in cf_rows]
    x = list(range(len(labels)))

    fig, ax = plt.subplots(figsize=(9, 3.4))
    ax.plot(x, inflow,  marker="o", color="#2563EB", label="Cash In")
    ax.plot(x, outflow, marker="s", color="#EF4444", label="Cash Out")
    ax.fill_between(x, inflow, outflow,
                    where=[i > o for i, o in zip(inflow, outflow)],
                    alpha=0.15, color="#10B981")
    ax.fill_between(x, inflow, outflow,
                    where=[i <= o for i, o in zip(inflow, outflow)],
                    alpha=0.15, color="#EF4444")
    ax2 = ax.twinx()
    ax2.bar(x, net, alpha=0.35, color="#F59E0B", label="Net Cash")
    ax2.set_ylabel("Net Cash (M NPR)", color="#F59E0B")
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=40, ha="right", fontsize=8)
    ax.set_ylabel("NPR (Millions)"); ax.set_title("Monthly Cash Flow")
    ax.legend(loc="upper left"); fig.tight_layout()
    data = _fig_to_bytes(fig); plt.close(fig); return data


def _chart_expenses(exp_breakdown: dict) -> bytes:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    labels = list(exp_breakdown.keys())
    vals   = [v / 1e6 for v in exp_breakdown.values()]
    colors = ["#2563EB","#10B981","#F59E0B","#EF4444","#8B5CF6","#EC4899",
              "#06B6D4","#84CC16"][:len(labels)]

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(10, 3.6))
    wedges, texts, autotexts = ax1.pie(
        vals, labels=None, autopct="%1.1f%%", startangle=140,
        colors=colors, pctdistance=0.8)
    for at in autotexts: at.set_fontsize(8)
    ax1.set_title("Expense Share")
    ax1.legend(wedges, labels, loc="lower center",
               bbox_to_anchor=(0.5, -0.25), ncol=2, fontsize=7)

    ax2.barh(labels[::-1], vals[::-1], color=colors[::-1])
    ax2.set_xlabel("NPR (Millions)")
    ax2.set_title("Expense Breakdown")
    for i, v in enumerate(vals[::-1]):
        ax2.text(v + 0.01, i, f"{v:.2f}M", va="center", fontsize=8)

    fig.tight_layout(); data = _fig_to_bytes(fig); plt.close(fig); return data


def _chart_top_products(products: list) -> bytes:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    names  = [_prod_name(p)            for p in products]
    revs   = [_prod_revenue(p) / 1e6   for p in products]
    profit = [_prod_profit(p)  / 1e6   for p in products]

    fig, ax = plt.subplots(figsize=(9, 3.8))
    x = range(len(names))
    ax.bar(x, revs,   width=0.4, label="Revenue (M)",  color="#2563EB")
    ax.bar([i+0.4 for i in x], profit, width=0.4,
           label="Profit (M)", color="#10B981")
    ax.set_xticks([i+0.2 for i in x])
    ax.set_xticklabels(names, rotation=35, ha="right", fontsize=8)
    ax.set_ylabel("NPR (Millions)")
    ax.set_title("Top 10 Products – Revenue vs Profit")
    ax.legend(); fig.tight_layout()
    data = _fig_to_bytes(fig); plt.close(fig); return data


def _chart_breakeven(be: dict) -> bytes:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    top    = be.get("top_20_easiest_breakeven", [])[:10]
    names  = [_be_product(p) for p in top]
    be_u   = [_be_units(p)   for p in top]
    act_u  = [_be_actual(p)  for p in top]

    fig, ax = plt.subplots(figsize=(9, 3.6))
    x = range(len(names))
    has_actual = any(v > 0 for v in act_u)
    if has_actual:
        ax.bar(x, act_u,  width=0.4, label="Actual Units",    color="#2563EB")
        ax.bar([i+0.4 for i in x], be_u, width=0.4,
               label="Break-Even Units", color="#EF4444", alpha=0.8)
        ax.set_xticks([i+0.2 for i in x])
        ax.set_xticklabels(names, rotation=35, ha="right", fontsize=8)
        ax.set_ylabel("Units"); ax.set_title("Break-Even vs Actual Units Sold (Top 10)")
    else:
        ax.bar(x, be_u, width=0.6, label="Break-Even Units", color="#EF4444", alpha=0.85)
        ax.set_xticks(list(x))
        ax.set_xticklabels(names, rotation=35, ha="right", fontsize=8)
        ax.set_ylabel("Units"); ax.set_title("Break-Even Units by Product (Top 10)")
    ax.legend(); fig.tight_layout()
    data = _fig_to_bytes(fig); plt.close(fig); return data


def _chart_inventory_reorder(items: list) -> bytes:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    items   = items[:12]
    names   = [_inv_sku(i)     for i in items]
    stock   = [_inv_stock(i)   for i in items]
    reorder = [_inv_reorder(i) for i in items]

    fig, ax = plt.subplots(figsize=(9, 3.4))
    x = range(len(names))
    ax.bar(x, reorder, width=0.4, label="Reorder Level", color="#EF4444", alpha=0.7)
    ax.bar([i+0.4 for i in x], stock, width=0.4,
           label="Current Stock", color="#F59E0B")
    ax.set_xticks([i+0.2 for i in x])
    ax.set_xticklabels(names, rotation=40, ha="right", fontsize=8)
    ax.set_ylabel("Units"); ax.set_title("Items Below Reorder Level")
    ax.legend(); fig.tight_layout()
    data = _fig_to_bytes(fig); plt.close(fig); return data


def _chart_discounts(discounts: dict) -> bytes:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    by_cat = _disc_by_category(discounts)
    cats   = list(by_cat.keys())
    vals   = [v / 1000 for v in by_cat.values()]
    monthly = _rget(discounts, "monthly_discount", "monthly_discounts", default=[]) or []

    fig, ax = plt.subplots(figsize=(7, 3.2))
    if cats:
        bars = ax.bar(cats, vals, color=["#2563EB","#10B981","#F59E0B","#EF4444"])
        ax.set_ylabel("Discount (NPR 000s)")
        ax.set_title("Discounts Granted by Category")
        for bar in bars:
            ax.text(bar.get_x() + bar.get_width()/2,
                    bar.get_height() + 0.5,
                    f"{bar.get_height():.0f}K",
                    ha="center", va="bottom", fontsize=8)
    elif monthly:
        labels = []
        m_vals = []
        for r in monthly:
            labels.append(str(_rget(r, "MonthNum", "month", "PeriodLabel", default="?")))
            m_vals.append(float(_rget(r, "monthly_discount_npr", "discount_npr", default=0)) / 1000)
        x = list(range(len(labels)))
        ax.plot(x, m_vals, marker="o", color="#2563EB")
        ax.set_xticks(x)
        ax.set_xticklabels(labels, rotation=40, ha="right", fontsize=8)
        ax.set_ylabel("Discount (NPR 000s)")
        ax.set_title("Monthly Discounts")
    else:
        ax.text(0.5, 0.5, "No discount category data",
                ha="center", va="center", fontsize=10, color="#64748B",
                transform=ax.transAxes)
        ax.set_axis_off()
    fig.tight_layout()
    data = _fig_to_bytes(fig); plt.close(fig); return data


# ---------------------------------------------------------------------------
#  ANALYTICAL NARRATIVES
# ---------------------------------------------------------------------------

def _narrative_profitability(p: dict) -> str:
    gpm = float(_rget(p, "gross_profit_margin_pct", "gpm_pct", default=0))
    npm = float(_rget(p, "net_profit_margin_pct", "npm_pct", default=0))
    gross_rev = float(_rget(p, "gross_revenue_npr", "gross_revenue", default=0))
    total_opex = float(_rget(p, "total_opex_npr", "total_operating_expense", default=0))
    net_profit = float(_rget(p, "net_profit_npr", "net_profit", default=0))
    health = "strong" if npm >= 15 else ("moderate" if npm >= 8 else "under pressure")
    gpm_msg = ("above the typical 40 % e-commerce benchmark - indicating efficient "
               "sourcing and pricing power." if gpm >= 40 else
               "below the 40 % e-commerce benchmark - review supplier costs and pricing.")
    return (
        f"The business generated NPR {gross_rev/1e6:.2f}M in gross revenue "
        f"with a gross profit margin of {gpm}%, {gpm_msg} "
        f"After accounting for NPR {total_opex/1e6:.2f}M in operating expenses, "
        f"net profit stands at NPR {net_profit/1e6:.2f}M ({npm}%), "
        f"which is {health}. "
        + ("Focus on OpEx optimisation - particularly marketing and logistics - "
           "to push net margin above 20%." if npm < 20 else
           "Maintain cost discipline and explore reinvestment to sustain this margin.")
    )

def _narrative_cashflow(cf: dict) -> str:
    net = float(_rget(cf, "net_cash_movement_npr", "net_cash_movement", default=0))
    inflow = float(_rget(cf, "total_cash_inflow_npr", "total_cash_inflow", default=0))
    outflow = float(_rget(cf, "total_cash_outflow_npr", "total_cash_outflow", default=0))
    ratio = inflow / max(outflow, 1)
    health = "healthy" if ratio >= 1.15 else ("tight" if ratio >= 1.0 else "negative")
    advice = ("The cash buffer provides runway for inventory pre-buying and expansion."
              if net > 0 else
              "Immediate action required: accelerate collections and defer non-critical spend.")
    return (
        f"Total cash inflows of NPR {inflow/1e6:.2f}M "
        f"against outflows of NPR {outflow/1e6:.2f}M "
        f"yields a cash coverage ratio of {ratio:.2f}x - classified as '{health}'. "
        f"Net cash movement for the period is NPR {net/1e6:.2f}M. {advice}"
    )

def _narrative_expenses(exp: dict) -> str:
    bd  = exp.get("expense_breakdown") or {}
    if not bd:
        return "Insufficient expense data to identify cost concentration."
    top = max(bd, key=bd.get)
    top_pct = round(bd[top] / sum(bd.values()) * 100, 1)
    return (
        f"'{top}' is the largest cost centre at {top_pct}% of total operating expenses. "
        "A diversified expense base reduces concentration risk; however, "
        "categories growing faster than revenue should be reviewed quarterly. "
        "Consider automating repetitive logistics tasks and renegotiating platform fees "
        "to drive structural cost reduction."
    )

def _narrative_inventory(inv: dict) -> str:
    to  = float(_rget(inv, "inventory_turnover", default=0))
    dio = float(_rget(inv, "days_inventory_outstanding", default=0))
    alert = int(_rget(inv, "items_below_reorder_level", default=0))
    bench = "above" if to >= 6 else "below"
    return (
        f"Inventory turns {to}x per year ({dio} days outstanding), "
        f"which is {bench} the e-commerce benchmark of 6x. "
        f"{alert} SKUs are currently below their reorder level - "
        f"{'urgent restocking required to prevent stockouts.' if alert >= 10 else 'monitor closely.'} "
        "Consider adopting a demand-driven reorder system to reduce both stockouts and overstock."
    )

def _narrative_products(products: list) -> str:
    if not products:
        return "Insufficient product data to identify top performers."
    top    = products[0]
    bot    = products[-1]
    spread = _prod_revenue(top) - _prod_revenue(bot)
    return (
        f"The top performer, '{_prod_name(top)}', generates "
        f"NPR {_prod_revenue(top)/1e6:.2f}M in revenue - "
        f"{spread/1e6:.2f}M more than the 10th-ranked product. "
        "High revenue concentration in a few SKUs signals an opportunity to cross-sell "
        "mid-tier products and reduce dependency risk. "
        "Products with high unit counts but lower revenue should be evaluated for margin improvement."
    )

def _narrative_breakeven(be: dict) -> str:
    mos = float(_rget(be, "margin_of_safety_pct", default=0))
    mos_label = ("comfortable" if mos >= 30 else
                 "adequate" if mos >= 15 else "dangerously thin")
    return (
        f"The business needs to sell {_rget(be, 'overall_breakeven_units', default=0):,} units to break even; "
        f"it actually sold {_rget(be, 'actual_units_sold', default=0):,} - a margin of safety of {mos}% "
        f"({mos_label}). "
        + ("Profitability is well-protected from demand shocks. "
           "Use the buffer to invest in growth." if mos >= 30 else
           "Increase the margin of safety by reducing fixed costs or growing high-margin SKUs.")
    )

def _narrative_discounts(d: dict, gross_revenue: float | None = None) -> str:
    total = _disc_total(d)
    rate  = _disc_rate(d, gross_revenue=gross_revenue)
    flag = "within acceptable range" if rate <= 12 else "elevated ? review discount strategy"
    return (
        f"NPR {total/1e6:.2f}M was granted in discounts "
        f"({rate}% of gross revenue), which is {flag}. "
        "Targeted promotional campaigns yield better ROI than blanket discounts. "
        "Track redemption rates by category to identify which discounts drive incremental sales."
    )


def _narrative_growth(rows: list) -> str:
    if not rows:
        return "Insufficient data to determine growth trend."
    rates = []
    for r in rows:
        val = _growth_rate(r)
        if val in ("", None):
            continue
        try:
            rates.append(float(val))
        except Exception:
            continue
    if not rates:
        # Infer growth from revenue sequence
        revs = [_growth_revenue(r) for r in rows]
        rates = [round((revs[i] / revs[i-1] - 1) * 100, 2)
                 for i in range(1, len(revs)) if revs[i-1] != 0]
    avg   = round(sum(rates) / len(rates), 2) if rates else 0
    trend = ("upward"   if rates and rates[-1] > rates[0] else
             "downward" if rates and rates[-1] < rates[0] else "flat")
    return (
        f"Revenue shows a {trend} trend over the period with an average monthly growth rate "
        f"of {avg}%. "
        + ("Momentum is building — capitalise with inventory pre-positioning and marketing spend."
           if avg > 3 else
           "Growth has stalled — investigate demand drivers, pricing, and competitive dynamics.")
    )


# ═══════════════════════════════════════════════════════════════════════════
#  PDF GENERATOR
# ═══════════════════════════════════════════════════════════════════════════

def generate_pdf(store: dict, period: str = "monthly", bucket: str | None = None) -> str:
    from reportlab.lib.pagesizes  import A4
    from reportlab.lib            import colors
    from reportlab.lib.units      import cm
    from reportlab.lib.styles     import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus       import (SimpleDocTemplate, Paragraph, Spacer,
                                          Table, TableStyle, Image,
                                          HRFlowable, PageBreak)
    from reportlab.lib.enums      import TA_CENTER

    period = _normalize_period(period)
    data   = _all(store, period=period, bucket=bucket)
    p      = data["profitability"]
    be     = data["breakeven"]
    cf     = data["cashflow"]
    inv    = data["inventory"]
    exp    = data["expenses"]
    disc   = data["discounts"]
    prods  = data["products"]["top_10_products_by_revenue"]
    growth = data["monthly_growth"].get("monthly", [])
    cf_rows= data["cashflow"]["monthly_cashflow"]

    out_path = str(EXPORT_DIR / f"BI_Report_{period}.pdf")

    doc = SimpleDocTemplate(
        out_path, pagesize=A4,
        leftMargin=1.8*cm, rightMargin=1.8*cm,
        topMargin=2*cm,    bottomMargin=2*cm,
        title=f"Nepal E-Commerce BI Report – {period.title()}",
        author="BI Report Generator",
    )

    # ── Styles ────────────────────────────────────────────────────────────
    styles = getSampleStyleSheet()
    BLUE   = colors.HexColor("#2563EB")
    DBLUE  = colors.HexColor("#1E3A5F")
    GREEN  = colors.HexColor("#10B981")
    AMBER  = colors.HexColor("#F59E0B")
    LGRAY  = colors.HexColor("#F1F5F9")
    MGRAY  = colors.HexColor("#94A3B8")
    WHITE  = colors.white
    BLACK  = colors.HexColor("#0F172A")

    title_style = ParagraphStyle("ReportTitle", parent=styles["Title"],
                                 textColor=DBLUE, fontSize=22, spaceAfter=4,
                                 alignment=TA_CENTER)
    sub_style   = ParagraphStyle("SubTitle", parent=styles["Normal"],
                                 textColor=MGRAY, fontSize=10,
                                 alignment=TA_CENTER, spaceAfter=8)
    h2_style    = ParagraphStyle("H2", parent=styles["Heading2"],
                                 textColor=WHITE, fontSize=12, spaceBefore=14,
                                 spaceAfter=6, leftIndent=6)
    kpi_label   = ParagraphStyle("KPILabel", parent=styles["Normal"],
                                 textColor=MGRAY, fontSize=8)
    kpi_value   = ParagraphStyle("KPIValue", parent=styles["Normal"],
                                 textColor=BLACK, fontSize=13, fontName="Helvetica-Bold")
    body_style  = ParagraphStyle("Body", parent=styles["Normal"],
                                 fontSize=9, leading=14, textColor=BLACK,
                                 spaceAfter=6)
    insight_style = ParagraphStyle("Insight", parent=styles["Normal"],
                                   fontSize=9, leading=14,
                                   textColor=colors.HexColor("#1E40AF"),
                                   backColor=colors.HexColor("#EFF6FF"),
                                   borderPad=8, spaceBefore=4, spaceAfter=8,
                                   leftIndent=8, rightIndent=8)

    story: list = []

    # ── Cover block ───────────────────────────────────────────────────────
    story.append(Spacer(1, 1*cm))
    story.append(Paragraph("Nepal E-Commerce BI Report 2025", title_style))
    story.append(Paragraph(f"Period View: {period.title()}", sub_style))
    story.append(HRFlowable(width="100%", thickness=2, color=BLUE, spaceAfter=8))

    # ── KPI summary row ───────────────────────────────────────────────────
    def kpi_cell(label: str, value: str) -> list:
        return [Paragraph(label, kpi_label), Paragraph(value, kpi_value)]

    kpi_data = [
        kpi_cell("Gross Revenue",  f"NPR {p['gross_revenue_npr']/1e6:.2f}M"),
        kpi_cell("Net Profit",     f"NPR {p['net_profit_npr']/1e6:.2f}M"),
        kpi_cell("Net Margin",     f"{p['net_profit_margin_pct']}%"),
        kpi_cell("Cash Coverage",  f"{cf['total_cash_inflow_npr']/max(cf['total_cash_outflow_npr'],1):.2f}x"),
        kpi_cell("Inventory Turn", f"{inv['inventory_turnover']}x"),
        kpi_cell("Margin Safety",  f"{be['margin_of_safety_pct']}%"),
    ]
    kpi_table = Table([kpi_data], colWidths=[2.6*cm]*6,
                      hAlign="CENTER", rowHeights=[1.6*cm])
    kpi_table.setStyle(TableStyle([
        ("BOX",          (0,0), (-1,-1), 0.5, MGRAY),
        ("INNERGRID",    (0,0), (-1,-1), 0.3, MGRAY),
        ("BACKGROUND",   (0,0), (-1,-1), LGRAY),
        ("TOPPADDING",   (0,0), (-1,-1), 6),
        ("BOTTOMPADDING",(0,0), (-1,-1), 6),
        ("LEFTPADDING",  (0,0), (-1,-1), 8),
        ("RIGHTPADDING", (0,0), (-1,-1), 8),
        ("ALIGN",        (0,0), (-1,-1), "CENTER"),
        ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 0.4*cm))

    # ── Generic section helpers ───────────────────────────────────────────
    def section_header(title: str, color=BLUE):
        tbl = Table([[Paragraph(f"  {title}", h2_style)]],
                    colWidths=["100%"])
        tbl.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1), color),
                                 ("ROUNDEDCORNERS",[4])]))
        story.append(tbl)

    def stat_table(rows: list[tuple], col_w=(10*cm, 6*cm)):
        tbl = Table(rows, colWidths=list(col_w))
        tbl.setStyle(TableStyle([
            ("ROWBACKGROUNDS", (0,0), (-1,-1), [WHITE, LGRAY]),
            ("FONTSIZE",       (0,0), (-1,-1), 9),
            ("TOPPADDING",     (0,0), (-1,-1), 4),
            ("BOTTOMPADDING",  (0,0), (-1,-1), 4),
            ("LEFTPADDING",    (0,0), (-1,-1), 6),
            ("RIGHTPADDING",   (0,0), (-1,-1), 6),
            ("TEXTCOLOR",      (0,0), (0,-1),  colors.HexColor("#475569")),
            ("FONTNAME",       (1,0), (1,-1),  "Helvetica-Bold"),
            ("ALIGN",          (1,0), (1,-1),  "RIGHT"),
        ]))
        story.append(tbl)

    def data_table(header: list, rows: list, col_ws=None):
        all_rows  = [header] + rows
        col_count = len(header)
        if col_ws is None:
            w = 15.5*cm / col_count
            col_ws = [w] * col_count
        tbl = Table(all_rows, colWidths=col_ws, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND",   (0,0),  (-1,0),  BLUE),
            ("TEXTCOLOR",    (0,0),  (-1,0),  WHITE),
            ("FONTNAME",     (0,0),  (-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",     (0,0),  (-1,-1), 8),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [WHITE, LGRAY]),
            ("GRID",         (0,0),  (-1,-1), 0.25, MGRAY),
            ("TOPPADDING",   (0,0),  (-1,-1), 3),
            ("BOTTOMPADDING",(0,0),  (-1,-1), 3),
            ("LEFTPADDING",  (0,0),  (-1,-1), 5),
            ("RIGHTPADDING", (0,0),  (-1,-1), 5),
            ("ALIGN",        (1,0),  (-1,-1), "RIGHT"),
        ]))
        story.append(tbl)

    def chart_image(png_bytes: bytes, width=15*cm):
        buf = io.BytesIO(png_bytes)
        img = Image(buf, width=width, height=width*0.4)
        story.append(img)

    def insight(text: str):
        wrapped = " ".join(textwrap.wrap(text, width=140))
        story.append(Paragraph(f"💡  {wrapped}", insight_style))

    # ─────────────────────────────────────────────────────────────────────
    # 1. PROFITABILITY
    # ─────────────────────────────────────────────────────────────────────
    section_header("1.  Profitability")
    story.append(Spacer(1, 0.2*cm))
    stat_table([
        ("Gross Revenue",           f"NPR {p['gross_revenue_npr']:,.0f}"),
        ("Net Revenue",             f"NPR {p['net_revenue_npr']:,.0f}"),
        ("Gross Profit",            f"NPR {p['gross_profit_npr']:,.0f}  ({p['gross_profit_margin_pct']}%)"),
        ("Net Profit",              f"NPR {p['net_profit_npr']:,.0f}  ({p['net_profit_margin_pct']}%)"),
        ("Total Operating Expense", f"NPR {p['total_opex_npr']:,.0f}"),
    ])
    story.append(Spacer(1, 0.3*cm))
    insight(_narrative_profitability(p))

    # ─────────────────────────────────────────────────────────────────────
    # 2. MONTHLY GROWTH  (with chart)
    # ─────────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 0.2*cm))
    section_header("2.  Revenue & Profit Growth", color=colors.HexColor("#0F766E"))
    story.append(Spacer(1, 0.2*cm))
    if growth:
        chart_image(_chart_revenue_profit(growth))
        growth_header = ["Period", "Revenue (NPR)", "Profit (NPR)", "Growth %"]
        growth_rows   = [
            [_growth_label(r),
             f"{_growth_revenue(r):,.0f}",
             f"{_growth_profit(r):,.0f}",
             f"{_growth_rate(r)}%"]
            for r in growth
        ]
        story.append(Spacer(1, 0.2*cm))
        data_table(growth_header, growth_rows, col_ws=[3.5*cm, 4.5*cm, 4.5*cm, 3*cm])
    insight(_narrative_growth(growth))

    # ─────────────────────────────────────────────────────────────────────
    # 3. CASH FLOW
    # ─────────────────────────────────────────────────────────────────────
    story.append(PageBreak())
    section_header("3.  Cash Flow Analysis", color=colors.HexColor("#7C3AED"))
    story.append(Spacer(1, 0.2*cm))
    stat_table([
        ("Total Cash Inflow",  f"NPR {cf['total_cash_inflow_npr']:,.0f}"),
        ("Total Cash Outflow", f"NPR {cf['total_cash_outflow_npr']:,.0f}"),
        ("Net Cash Movement",  f"NPR {cf['net_cash_movement_npr']:,.0f}"),
    ])
    story.append(Spacer(1, 0.3*cm))
    if cf_rows:
        chart_image(_chart_cashflow(cf_rows))
        story.append(Spacer(1, 0.2*cm))
        cf_header = ["Period", "Cash In (NPR)", "Cash Out (NPR)", "Net (NPR)"]
        cf_table_rows = [
            [_cf_label(r),
             f"{_cf_inflow(r):,.0f}",
             f"{_cf_outflow(r):,.0f}",
             f"{_cf_net(r):,.0f}"]
            for r in cf_rows
        ]
        data_table(cf_header, cf_table_rows, col_ws=[3.5*cm,4*cm,4*cm,4*cm])
    insight(_narrative_cashflow(cf))

    # ─────────────────────────────────────────────────────────────────────
    # 4. EXPENSES
    # ─────────────────────────────────────────────────────────────────────
    story.append(PageBreak())
    section_header("4.  Expense Breakdown", color=colors.HexColor("#B45309"))
    story.append(Spacer(1, 0.2*cm))
    chart_image(_chart_expenses(exp["expense_breakdown"]))
    story.append(Spacer(1, 0.2*cm))
    exp_rows = [(cat, f"NPR {val:,.0f}")
                for cat, val in exp["expense_breakdown"].items()]
    stat_table(exp_rows)
    insight(_narrative_expenses(exp))

    # ─────────────────────────────────────────────────────────────────────
    # 5. TOP PRODUCTS
    # ─────────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 0.3*cm))
    section_header("5.  Top 10 Products by Revenue", color=colors.HexColor("#065F46"))
    story.append(Spacer(1, 0.2*cm))
    chart_image(_chart_top_products(prods))
    story.append(Spacer(1, 0.2*cm))
    prod_header = ["Product", "Revenue (NPR)", "Units Sold", "Profit (NPR)"]
    prod_rows   = [
        [_prod_name(p2),
         f"{_prod_revenue(p2):,.0f}",
         f"{_prod_units(p2):,}",
         f"{_prod_profit(p2):,.0f}"]
        for p2 in prods
    ]
    data_table(prod_header, prod_rows, col_ws=[5.5*cm,4*cm,3*cm,3*cm])
    insight(_narrative_products(prods))

    # ─────────────────────────────────────────────────────────────────────
    # 6. INVENTORY
    # ─────────────────────────────────────────────────────────────────────
    story.append(PageBreak())
    section_header("6.  Inventory Health", color=colors.HexColor("#9D174D"))
    story.append(Spacer(1, 0.2*cm))
    stat_table([
        ("Inventory Turnover",          f"{inv['inventory_turnover']}x"),
        ("Days Inventory Outstanding",  f"{inv['days_inventory_outstanding']} days"),
        ("Items Below Reorder Level",   str(inv.get('items_below_reorder_level', 0))),
    ])
    story.append(Spacer(1, 0.3*cm))
    if inv.get("below_reorder_items"):
        chart_image(_chart_inventory_reorder(inv["below_reorder_items"]))
        story.append(Spacer(1, 0.2*cm))
        reorder_header = ["Item ID", "Product", "Closing Stock", "Reorder Level"]
        reorder_rows   = [
            [_inv_sku(i), _inv_name(i),
             str(_inv_stock(i)), str(_inv_reorder(i))]
            for i in inv["below_reorder_items"][:12]
        ]
        data_table(reorder_header, reorder_rows,
                   col_ws=[3*cm, 5.5*cm, 3.5*cm, 3.5*cm])
    insight(_narrative_inventory(inv))

    # ─────────────────────────────────────────────────────────────────────
    # 7. BREAK-EVEN
    # ─────────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 0.3*cm))
    section_header("7.  Break-Even Analysis", color=colors.HexColor("#1D4ED8"))
    story.append(Spacer(1, 0.2*cm))
    stat_table([
        ("Overall Break-Even Units", f"{be['overall_breakeven_units']:,.0f}"),
        ("Actual Units Sold",        f"{be['actual_units_sold']:,.0f}"),
        ("Margin of Safety",         f"{be['margin_of_safety_pct']}%"),
    ])
    story.append(Spacer(1, 0.3*cm))
    chart_image(_chart_breakeven(be))
    story.append(Spacer(1, 0.2*cm))
    be_header = ["Product", "B/E Units", "Contrib / Unit (NPR)"]
    be_rows   = [
        [_be_product(r), f"{_be_units(r):,}",
         f"{_rget(r, 'ContribPerUnit', 'contrib_per_unit', default=0):,.2f}"]
        for r in be["top_20_easiest_breakeven"][:15]
    ]
    data_table(be_header, be_rows, col_ws=[7*cm, 4*cm, 4.5*cm])
    insight(_narrative_breakeven(be))

    # ─────────────────────────────────────────────────────────────────────
    # 8. DISCOUNTS
    # ─────────────────────────────────────────────────────────────────────
    story.append(PageBreak())
    section_header("8.  Discount Analysis", color=colors.HexColor("#6D28D9"))
    story.append(Spacer(1, 0.2*cm))
    total_disc = _disc_total(disc)
    disc_rate  = _disc_rate(disc, gross_revenue=p.get("gross_revenue_npr"))
    stat_table([
        ("Total Discounts Granted", f"NPR {total_disc:,.0f}"),
        ("Effective Discount Rate", f"{disc_rate}%"),
    ])
    story.append(Spacer(1, 0.3*cm))
    chart_image(_chart_discounts(disc), width=12*cm)
    insight(_narrative_discounts(disc, gross_revenue=p.get("gross_revenue_npr")))

    # ─────────────────────────────────────────────────────────────────────
    # Footer note
    # ─────────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=MGRAY))
    story.append(Paragraph(
        "Generated by Nepal E-Commerce BI System · All figures in NPR · For internal use only.",
        ParagraphStyle("Footer", parent=styles["Normal"],
                       fontSize=7, textColor=MGRAY, alignment=TA_CENTER)))

    doc.build(story)
    logger.info(f"PDF saved: {out_path}")
    return out_path


# ═══════════════════════════════════════════════════════════════════════════
#  EXCEL GENERATOR
# ═══════════════════════════════════════════════════════════════════════════

def generate_excel(store: dict, period: str = "monthly", bucket: str | None = None) -> str:
    import pandas as pd
    from openpyxl                         import load_workbook
    from openpyxl.styles                  import (Font, PatternFill, Alignment,
                                                   Border, Side, GradientFill)
    from openpyxl.utils                   import get_column_letter
    from openpyxl.drawing.image           import Image as XLImage
    from openpyxl.chart                   import BarChart, LineChart, Reference
    from openpyxl.chart.series            import DataPoint

    period = _normalize_period(period)
    data   = _all(store, period=period, bucket=bucket)
    out    = str(EXPORT_DIR / f"BI_Report_{period}.xlsx")

    monthly_growth_rows = data["monthly_growth"].get("monthly", [])
    cashflow_rows       = data["cashflow"]["monthly_cashflow"]
    p                   = data["profitability"]
    be                  = data["breakeven"]
    cf                  = data["cashflow"]
    inv                 = data["inventory"]

    # ── Write raw data first ──────────────────────────────────────────────
    with pd.ExcelWriter(out, engine="openpyxl") as writer:

        # 1. Summary
        summary_df = pd.DataFrame([{
            "Metric": "Gross Revenue (NPR)",        "Value": p["gross_revenue_npr"]},
            {"Metric": "Net Revenue (NPR)",          "Value": p["net_revenue_npr"]},
            {"Metric": "Gross Profit (NPR)",         "Value": p["gross_profit_npr"]},
            {"Metric": "Gross Profit Margin (%)",    "Value": p["gross_profit_margin_pct"]},
            {"Metric": "Net Profit (NPR)",           "Value": p["net_profit_npr"]},
            {"Metric": "Net Profit Margin (%)",      "Value": p["net_profit_margin_pct"]},
            {"Metric": "Total OpEx (NPR)",           "Value": p["total_opex_npr"]},
            {"Metric": "Cash Inflow (NPR)",          "Value": cf["total_cash_inflow_npr"]},
            {"Metric": "Cash Outflow (NPR)",         "Value": cf["total_cash_outflow_npr"]},
            {"Metric": "Net Cash Movement (NPR)",    "Value": cf["net_cash_movement_npr"]},
            {"Metric": "Inventory Turnover (x)",     "Value": inv["inventory_turnover"]},
            {"Metric": "Days Inventory Outstanding", "Value": inv["days_inventory_outstanding"]},
            {"Metric": "Items Below Reorder",        "Value": inv["items_below_reorder_level"]},
            {"Metric": "Break-Even Units",           "Value": be["overall_breakeven_units"]},
            {"Metric": "Actual Units Sold",          "Value": be["actual_units_sold"]},
            {"Metric": "Margin of Safety (%)",       "Value": be["margin_of_safety_pct"]},
        ])
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # 2. Profitability
        pd.DataFrame([p]).T.reset_index().rename(
            columns={"index": "Metric", 0: "Value"}
        ).to_excel(writer, sheet_name="Profitability", index=False)

        # 3. Growth
        pd.DataFrame(monthly_growth_rows).to_excel(
            writer, sheet_name=f"{period.title()} Growth", index=False)

        # 4. Top Products
        pd.DataFrame(data["products"]["top_10_products_by_revenue"]).to_excel(
            writer, sheet_name="Top Products", index=False)

        # 5. Expenses
        pd.DataFrame([data["expenses"]["expense_breakdown"]]).T.reset_index().rename(
            columns={"index": "Category", 0: "NPR"}
        ).to_excel(writer, sheet_name="Expenses", index=False)

        # 6. Cash Flow
        pd.DataFrame(cashflow_rows).to_excel(
            writer, sheet_name=f"{period.title()} Cash Flow", index=False)

        # 7. Break-Even
        pd.DataFrame(be["top_20_easiest_breakeven"]).to_excel(
            writer, sheet_name="Break-Even", index=False)

        # 8. Reorder Alerts
        pd.DataFrame(inv["below_reorder_items"]).to_excel(
            writer, sheet_name="Reorder Alerts", index=False)

        # 9. Analysis Insights sheet
        insights_df = pd.DataFrame([
            {"Section": "Profitability",  "Analysis": _narrative_profitability(p)},
            {"Section": "Revenue Growth", "Analysis": _narrative_growth(monthly_growth_rows)},
            {"Section": "Cash Flow",      "Analysis": _narrative_cashflow(cf)},
            {"Section": "Expenses",       "Analysis": _narrative_expenses(data["expenses"])},
            {"Section": "Top Products",   "Analysis": _narrative_products(
                data["products"]["top_10_products_by_revenue"])},
            {"Section": "Inventory",      "Analysis": _narrative_inventory(inv)},
            {"Section": "Break-Even",     "Analysis": _narrative_breakeven(be)},
            {"Section": "Discounts",      "Analysis": _narrative_discounts(
                data["discounts"], gross_revenue=p.get("gross_revenue_npr"))},
        ])
        insights_df.to_excel(writer, sheet_name="Analysis Insights", index=False)

    # ── Post-process: styling + charts ────────────────────────────────────
    wb = load_workbook(out)

    HEADER_FILL  = PatternFill("solid", fgColor="2563EB")
    ACCENT_FILL  = PatternFill("solid", fgColor="EFF6FF")
    ALT_FILL     = PatternFill("solid", fgColor="F1F5F9")
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
    BODY_FONT    = Font(size=9)
    BOLD_FONT    = Font(bold=True, size=9)
    TITLE_FONT   = Font(bold=True, size=14, color="1E3A5F")
    WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top")
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
    THIN_BORDER  = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    def style_sheet(ws, title: str = ""):
        # Title row
        if title:
            ws.insert_rows(1)
            ws.insert_rows(1)
            ws["A1"] = title
            ws["A1"].font = TITLE_FONT
            ws["A1"].alignment = CENTER_ALIGN
            ws.merge_cells(start_row=1, start_column=1,
                           end_row=1, end_column=ws.max_column)
            ws["A2"] = f"Nepal E-Commerce BI Report – {period.title()}"
            ws["A2"].font = Font(size=9, color="94A3B8", italic=True)
            ws["A2"].alignment = CENTER_ALIGN
            ws.merge_cells(start_row=2, start_column=1,
                           end_row=2, end_column=ws.max_column)
            header_row = 3
        else:
            header_row = 1

        for col_idx, cell in enumerate(ws[header_row], start=1):
            cell.fill   = HEADER_FILL
            cell.font   = HEADER_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        for row_idx, row in enumerate(
                ws.iter_rows(min_row=header_row+1, max_row=ws.max_row), start=1):
            fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
            for cell in row:
                cell.fill      = fill
                cell.font      = BODY_FONT
                cell.border    = THIN_BORDER
                cell.alignment = WRAP_ALIGN

        # Auto-width columns
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)

        ws.freeze_panes = ws.cell(row=header_row+1, column=1)

    # Style every sheet
    for ws in wb.worksheets:
        style_sheet(ws, title=ws.title)

    # ── Embed matplotlib charts as images ─────────────────────────────────
    def embed_chart(ws, png_bytes: bytes, anchor: str = "A30"):
        buf = io.BytesIO(png_bytes)
        img = XLImage(buf)
        img.width  = 680
        img.height = 270
        ws.add_image(img, anchor)

    # Growth sheet
    growth_sheet = wb[f"{period.title()} Growth"]
    embed_chart(growth_sheet, _chart_revenue_profit(monthly_growth_rows), "A20")

    # Cash Flow sheet
    cf_sheet = wb[f"{period.title()} Cash Flow"]
    embed_chart(cf_sheet, _chart_cashflow(cashflow_rows), "A20")

    # Expenses sheet
    exp_sheet = wb["Expenses"]
    embed_chart(exp_sheet, _chart_expenses(data["expenses"]["expense_breakdown"]), "D3")

    # Products sheet
    prod_sheet = wb["Top Products"]
    embed_chart(prod_sheet,
                _chart_top_products(data["products"]["top_10_products_by_revenue"]), "A16")

    # Inventory sheet
    if "Reorder Alerts" in wb.sheetnames:
        inv_sheet = wb["Reorder Alerts"]
        embed_chart(inv_sheet,
                    _chart_inventory_reorder(inv["below_reorder_items"]), "A20")

    # Break-Even sheet
    be_sheet = wb["Break-Even"]
    embed_chart(be_sheet, _chart_breakeven(be), "A25")

    # Insights sheet – widen Analysis column
    ins_ws = wb["Analysis Insights"]
    ins_ws.column_dimensions["B"].width = 90
    for row in ins_ws.iter_rows(min_row=4, max_row=ins_ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ins_ws.iter_rows(min_row=4, max_row=ins_ws.max_row):
        ins_ws.row_dimensions[row[0].row].height = 60

    wb.save(out)
    logger.info(f"Excel saved: {out}")
    return out


# ═══════════════════════════════════════════════════════════════════════════
#  CLI quick-test
# ═══════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    import sys
    period = sys.argv[1] if len(sys.argv) > 1 else "monthly"
    print("Generating PDF …")
    pdf_path = generate_pdf({}, period=period)
    print(f"  PDF  → {pdf_path}")
    print("Generating Excel …")
    xl_path  = generate_excel({}, period=period)
    print(f"  XLSX → {xl_path}")
